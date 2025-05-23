"""
Export manager for the Dark Web Discovery System.
Provides functionality to export data in multiple formats.
"""

import os
import csv
import json
import datetime
import logging
import tempfile
from typing import Dict, List, Any, Optional, Union, Tuple

import pandas as pd
import networkx as nx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import Config

# Configure logger
def log_action(message):
    """Log actions with timestamp."""
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")
    logging.info(message)


class ExportManager:
    """
    Manages data exports in various formats.
    Provides customization options and templates.
    """
    
    # Export formats
    FORMATS = {
        "csv": "CSV (Comma Separated Values)",
        "excel": "Excel Spreadsheet",
        "json": "JSON (JavaScript Object Notation)",
        "graphml": "GraphML (Graph Markup Language)",
        "html": "HTML Table"
    }
    
    # Export templates
    TEMPLATES = {
        "basic": {
            "name": "Basic",
            "description": "Basic export with minimal fields",
            "fields": ["url", "title", "category", "status", "last_checked"]
        },
        "detailed": {
            "name": "Detailed",
            "description": "Detailed export with all fields",
            "fields": ["url", "title", "description", "category", "status", "last_checked", 
                     "discovery_date", "discovery_source", "content_preview", "metadata"]
        },
        "network": {
            "name": "Network Analysis",
            "description": "Export optimized for network analysis",
            "fields": ["url", "title", "category", "status", "domain", "discovery_source"]
        },
        "safety": {
            "name": "Safety Analysis",
            "description": "Export focused on content safety",
            "fields": ["url", "title", "category", "status", "safety_score", "safety_categories", 
                     "was_filtered", "filter_reason"]
        }
    }
    
    def __init__(self, link_db, network_visualizer=None, export_dir: Optional[str] = None):
        """
        Initialize the export manager.
        
        Args:
            link_db: Database instance for accessing link data
            network_visualizer: Network visualizer instance for graph exports
            export_dir (str): Directory for saving exports
        """
        self.link_db = link_db
        self.network_visualizer = network_visualizer
        self.export_dir = export_dir or Config.EXPORT_DIR
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_links(self, 
                    format: str,
                    template: str = "basic",
                    filename: Optional[str] = None,
                    filters: Optional[Dict] = None,
                    custom_fields: Optional[List[str]] = None) -> str:
        """
        Export links data in the specified format.
        
        Args:
            format (str): Export format (csv, excel, json, graphml, html)
            template (str): Export template name or "custom"
            filename (str): Output filename (without extension)
            filters (dict): Filters to apply to the data
            custom_fields (list): Custom fields to include (for custom template)
            
        Returns:
            str: Path to the exported file
        """
        # Validate format
        if format not in self.FORMATS:
            raise ValueError(f"Invalid export format: {format}")
        
        # Get export fields
        if template == "custom" and custom_fields:
            fields = custom_fields
        elif template in self.TEMPLATES:
            fields = self.TEMPLATES[template]["fields"]
        else:
            fields = self.TEMPLATES["basic"]["fields"]
        
        # Get links data with filters
        filters = filters or {}
        links = self._get_filtered_links(filters)
        
        # Generate filename if not provided
        if not filename:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"dark_web_links_{template}_{timestamp}"
        
        # Call appropriate export method
        if format == "csv":
            return self._export_to_csv(links, fields, filename)
        elif format == "excel":
            return self._export_to_excel(links, fields, filename)
        elif format == "json":
            return self._export_to_json(links, fields, filename)
        elif format == "graphml":
            return self._export_to_graphml(links, filename)
        elif format == "html":
            return self._export_to_html(links, fields, filename)
        else:
            raise ValueError(f"Format {format} is recognized but not implemented")
    
    def _get_filtered_links(self, filters: Dict) -> List[Dict]:
        """
        Get links data with filters applied.
        
        Args:
            filters (dict): Filters to apply
            
        Returns:
            list: Filtered links data
        """
        # Apply filters to database query
        category = filters.get("category")
        status = filters.get("status")
        search_query = filters.get("search_query")
        include_blacklisted = filters.get("include_blacklisted", False)
        max_days_old = filters.get("max_days_old")
        limit = filters.get("limit", 1000)  # Default limit for exports
        
        # Get links from database
        links = self.link_db.get_links_with_filters(
            category=category,
            status=status,
            search_query=search_query,
            include_blacklisted=include_blacklisted,
            max_days_old=max_days_old,
            limit=limit
        )
        
        return links
    
    def _prepare_export_data(self, links: List[Dict], fields: List[str]) -> List[Dict]:
        """
        Prepare data for export by extracting required fields.
        
        Args:
            links (list): Links data
            fields (list): Fields to include
            
        Returns:
            list: Prepared data
        """
        export_data = []
        
        for link in links:
            # Extract metadata fields if needed
            if "metadata" in link:
                metadata = link["metadata"] or {}
                for key, value in metadata.items():
                    if key not in link:
                        link[key] = value
            
            # Extract only required fields
            item = {}
            for field in fields:
                if field in link:
                    # Handle special field formatting
                    if field == "metadata" and link[field]:
                        item[field] = json.dumps(link[field])
                    else:
                        item[field] = link[field]
                else:
                    item[field] = ""
            
            export_data.append(item)
        
        return export_data
    
    def _export_to_csv(self, links: List[Dict], fields: List[str], filename: str) -> str:
        """
        Export data to CSV format.
        
        Args:
            links (list): Links data
            fields (list): Fields to include
            filename (str): Output filename
            
        Returns:
            str: Path to the exported file
        """
        # Prepare data
        export_data = self._prepare_export_data(links, fields)
        
        # Generate output path
        if not filename.endswith(".csv"):
            filename += ".csv"
        output_path = os.path.join(self.export_dir, filename)
        
        # Write to CSV
        with open(output_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fields)
            writer.writeheader()
            writer.writerows(export_data)
        
        log_action(f"Exported {len(export_data)} links to CSV: {output_path}")
        return output_path
    
    def _export_to_excel(self, links: List[Dict], fields: List[str], filename: str) -> str:
        """
        Export data to Excel format with formatting.
        
        Args:
            links (list): Links data
            fields (list): Fields to include
            filename (str): Output filename
            
        Returns:
            str: Path to the exported file
        """
        # Prepare data
        export_data = self._prepare_export_data(links, fields)
        
        # Generate output path
        if not filename.endswith(".xlsx"):
            filename += ".xlsx"
        output_path = os.path.join(self.export_dir, filename)
        
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Write data
            df.to_excel(writer, sheet_name="Dark Web Links", index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets["Dark Web Links"]
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            
            # Apply formatting to header
            for col_num, column_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Adjust column width
                column_letter = get_column_letter(col_num)
                column_width = max(len(str(column_name)) + 2, 15)
                worksheet.column_dimensions[column_letter].width = column_width
            
            # Add metadata
            metadata_sheet = workbook.create_sheet("Export Metadata")
            
            metadata = [
                ["Export Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                ["Total Records", len(export_data)],
                ["Fields Included", ", ".join(fields)],
                ["Generated By", "Dark Web Discovery System"]
            ]
            
            for row_num, (key, value) in enumerate(metadata, 1):
                metadata_sheet.cell(row=row_num, column=1, value=key).font = Font(bold=True)
                metadata_sheet.cell(row=row_num, column=2, value=value)
                
            # Adjust column width
            metadata_sheet.column_dimensions["A"].width = 20
            metadata_sheet.column_dimensions["B"].width = 50
        
        log_action(f"Exported {len(export_data)} links to Excel: {output_path}")
        return output_path
    
    def _export_to_json(self, links: List[Dict], fields: List[str], filename: str) -> str:
        """
        Export data to JSON format.
        
        Args:
            links (list): Links data
            fields (list): Fields to include
            filename (str): Output filename
            
        Returns:
            str: Path to the exported file
        """
        # Prepare data
        export_data = self._prepare_export_data(links, fields)
        
        # Generate output path
        if not filename.endswith(".json"):
            filename += ".json"
        output_path = os.path.join(self.export_dir, filename)
        
        # Add metadata
        full_export = {
            "metadata": {
                "export_date": datetime.datetime.now().isoformat(),
                "record_count": len(export_data),
                "fields": fields,
                "generator": "Dark Web Discovery System"
            },
            "data": export_data
        }
        
        # Write to JSON
        with open(output_path, "w", encoding="utf-8") as jsonfile:
            json.dump(full_export, jsonfile, indent=2, ensure_ascii=False)
        
        log_action(f"Exported {len(export_data)} links to JSON: {output_path}")
        return output_path
    
    def _export_to_graphml(self, links: List[Dict], filename: str) -> str:
        """
        Export data to GraphML format for network analysis tools.
        
        Args:
            links (list): Links data
            filename (str): Output filename
            
        Returns:
            str: Path to the exported file
        """
        # Check if network visualizer is available
        if self.network_visualizer:
            # Build graph using network visualizer
            G = self.network_visualizer.build_network_graph()
            
            # Generate output path
            if not filename.endswith(".graphml"):
                filename += ".graphml"
            output_path = os.path.join(self.export_dir, filename)
            
            # Export to GraphML
            nx.write_graphml(G, output_path)
            
            log_action(f"Exported network graph with {len(G.nodes())} nodes to GraphML: {output_path}")
            return output_path
        else:
            # Create graph manually
            G = nx.DiGraph()
            
            # Add nodes
            for link in links:
                url = link.get("url", "")
                if url:
                    # Node attributes
                    attrs = {}
                    for key, value in link.items():
                        if key != "metadata" and value is not None:
                            attrs[key] = str(value)
                    
                    # Add metadata as attributes if available
                    if "metadata" in link and link["metadata"]:
                        for meta_key, meta_value in link["metadata"].items():
                            if meta_value is not None:
                                attrs[f"meta_{meta_key}"] = str(meta_value)
                    
                    # Add node
                    G.add_node(url, **attrs)
            
            # Add edges based on discovery sources
            for link in links:
                url = link.get("url", "")
                discovery_source = link.get("discovery_source", "")
                
                if url and discovery_source and "://" in discovery_source:
                    # Discovery source is a URL
                    if discovery_source in G.nodes():
                        G.add_edge(discovery_source, url, type="discovered_by")
            
            # Generate output path
            if not filename.endswith(".graphml"):
                filename += ".graphml"
            output_path = os.path.join(self.export_dir, filename)
            
            # Export to GraphML
            nx.write_graphml(G, output_path)
            
            log_action(f"Exported network graph with {len(G.nodes())} nodes to GraphML: {output_path}")
            return output_path
    
    def _export_to_html(self, links: List[Dict], fields: List[str], filename: str) -> str:
        """
        Export data to HTML table format.
        
        Args:
            links (list): Links data
            fields (list): Fields to include
            filename (str): Output filename
            
        Returns:
            str: Path to the exported file
        """
        # Prepare data
        export_data = self._prepare_export_data(links, fields)
        
        # Generate output path
        if not filename.endswith(".html"):
            filename += ".html"
        output_path = os.path.join(self.export_dir, filename)
        
        # Create DataFrame
        df = pd.DataFrame(export_data)
        
        # Generate HTML with styling
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>Dark Web Links Export</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333366; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th {{ background-color: #366092; color: white; text-align: left; padding: 8px; }}
                td {{ border: 1px solid #ddd; padding: 8px; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .metadata {{ background-color: #eef; padding: 10px; margin-bottom: 20px; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <h1>Dark Web Links Export</h1>
            
            <div class="metadata">
                <p><strong>Export Date:</strong> {datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
                <p><strong>Total Records:</strong> {len(export_data)}</p>
                <p><strong>Fields Included:</strong> {", ".join(fields)}</p>
                <p><strong>Generated By:</strong> Dark Web Discovery System</p>
            </div>
            
            {df.to_html(index=False, escape=True, classes="table")}
            
            <p style="margin-top: 20px; color: #666; font-size: 0.8em;">
                This export was generated by the Dark Web Discovery System. Data may be sensitive.
            </p>
        </body>
        </html>
        """
        
        # Write to HTML file
        with open(output_path, "w", encoding="utf-8") as htmlfile:
            htmlfile.write(html_content)
        
        log_action(f"Exported {len(export_data)} links to HTML: {output_path}")
        return output_path
    
    def get_available_formats(self) -> Dict[str, str]:
        """
        Get available export formats.
        
        Returns:
            dict: Format ID to description mapping
        """
        return self.FORMATS
    
    def get_available_templates(self) -> Dict[str, Dict]:
        """
        Get available export templates.
        
        Returns:
            dict: Template ID to template details mapping
        """
        return self.TEMPLATES
    
    def schedule_export(self, 
                       format: str,
                       template: str = "basic",
                       filters: Optional[Dict] = None,
                       schedule: str = "daily",
                       custom_fields: Optional[List[str]] = None) -> Dict:
        """
        Schedule a recurring export.
        
        Args:
            format (str): Export format
            template (str): Export template name or "custom"
            filters (dict): Filters to apply to the data
            schedule (str): Schedule frequency (daily, weekly, monthly)
            custom_fields (list): Custom fields to include (for custom template)
            
        Returns:
            dict: Scheduled export details
        """
        # Generate unique ID for the scheduled export
        export_id = f"export_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Create schedule details
        scheduled_export = {
            "id": export_id,
            "format": format,
            "template": template,
            "filters": filters or {},
            "schedule": schedule,
            "custom_fields": custom_fields,
            "created_at": datetime.datetime.now().isoformat(),
            "last_run": None,
            "next_run": self._calculate_next_run(schedule)
        }
        
        # Store scheduled export (implementation depends on system)
        # This is a placeholder - actual implementation would store to database
        log_action(f"Scheduled export created: {export_id} ({schedule} {format} export)")
        
        return scheduled_export
    
    def _calculate_next_run(self, schedule: str) -> str:
        """
        Calculate next run time based on schedule.
        
        Args:
            schedule (str): Schedule frequency
            
        Returns:
            str: Next run time as ISO format
        """
        now = datetime.datetime.now()
        
        if schedule == "daily":
            next_run = now.replace(hour=0, minute=0, second=0, microsecond=0) + datetime.timedelta(days=1)
        elif schedule == "weekly":
            # Next Monday
            days_ahead = 0 - now.weekday() + 7
            next_run = now.replace(hour=0, minute=0, second=0, microsecond=0) + datetime.timedelta(days=days_ahead)
        elif schedule == "monthly":
            # First day of next month
            if now.month == 12:
                next_run = now.replace(year=now.year+1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            else:
                next_run = now.replace(month=now.month+1, day=1, hour=0, minute=0, second=0, microsecond=0)
        else:
            # Default to daily
            next_run = now.replace(hour=0, minute=0, second=0, microsecond=0) + datetime.timedelta(days=1)
        
        return next_run.isoformat()
